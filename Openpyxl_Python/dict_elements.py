import openpyxl as xl
wb=xl.load_workbook("C://MyDataSets//student_openpyxl.xlsx")
sheet1=wb.active
d={}
l=[]

# for i in range(1,4):
#     # for j in range(1,8):
#     d.setdefault(sheet1.cell(1,i).value)
#     # d.setdefault(sheet1.cell(1,j).value,sheet1.cell(i,j).value)
for i in range(1,4):
    l=[]
    for j in range(2,7):
        l.append(sheet1.cell(j,i).value)
    d[(sheet1.cell(1,i).value)]=l
print(l)
print(d)

