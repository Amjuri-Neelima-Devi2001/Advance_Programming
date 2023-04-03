import openpyxl as xl
def listOfDict(wb,wb1):
    d=[]
    for i in range(1,sheet1.max_row):
        l={}
        for j in range(1,4):
            l.setdefault(sheet1.cell(1,j).value,sheet1.cell(i+1,j).value)
        d.append(l)
    return d

wb=xl.load_workbook("C://MyDataSets//daily_data.xlsx")
wb1=xl.load_workbook("C://MyDataSets//master_data.xlsx")
sheet1=wb.active
sheet2=wb1.active
x=listOfDict(wb,wb1)
print(x)
total=0
for i in x:
    total+=i['todays purchase']
print(total)
for i in x:
    ID=
    if i['id']==ID:

