import openpyxl
wb=openpyxl.load_workbook("C://MyDataSets//student_openpyxl.xlsx")
print(wb.sheetnames)
sheet=wb.active
print(sheet.max_row)
print(sheet.max_column)
# sheet["A4"]=
