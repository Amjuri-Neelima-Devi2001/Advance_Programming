import openpyxl as xl
def listOfDict(wb):
    d=[]
    for i in range(1,sheet1.max_row):
        l={}
        for j in range(1,4):
            l.setdefault(sheet1.cell(1,j).value,sheet1.cell(i+1,j).value)
        d.append(l)
    return d

wb=xl.load_workbook("C://MyDataSets//daily_data.xlsx")
sheet1=wb.active
x=listOfDict(wb)
print(x)
total=0
for i in x:
    total+=i['todays purchase']
print(total)


